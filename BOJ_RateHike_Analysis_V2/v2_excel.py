#!/usr/bin/env python3
"""
V2 Excel: Create comprehensive Excel with Input, Macro, EventReturns,
Summary_Financial, Summary_NonFinancial sheets.
"""
import pandas as pd
import numpy as np
import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Load data
df = pd.read_pickle('/home/ubuntu/boj_analysis/v2_full_results.pkl')
fin_rank = pd.read_pickle('/home/ubuntu/boj_analysis/v2_fin_ranking.pkl')
nonfin_rank = pd.read_pickle('/home/ubuntu/boj_analysis/v2_nonfin_ranking.pkl')
topix_ext = pd.read_pickle('/home/ubuntu/boj_analysis/topix_extended.pkl')
macro_data = pd.read_pickle('/home/ubuntu/boj_analysis/macro_data.pkl')

with open('/home/ubuntu/boj_analysis/v2_events_info.pkl', 'rb') as f:
    events = pickle.load(f)

# Styles
hf = Font(name='Meiryo', bold=True, size=9, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sf = Font(name='Meiryo', bold=True, size=9)
sfill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
nf = Font(name='Meiryo', size=9)
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

def write_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = hf; c.fill = hfill; c.alignment = ca; c.border = tb
    return c

def write_cell(ws, row, col, value, fmt=None, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = nf; c.border = tb; c.alignment = Alignment(horizontal='center')
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    return c

def color_val(val, neg_thresh=-1, pos_thresh=1):
    if pd.isna(val): return None
    if val <= neg_thresh: return red_fill
    if val >= pos_thresh: return green_fill
    return None

wb = Workbook()

# ==========================================
# Sheet 1: Input
# ==========================================
ws = wb.active
ws.title = 'Input'
ws.merge_cells('A1:H1')
ws['A1'] = '日銀利上げイベント 大型株耐性分析 V2（回帰調整版）'
ws['A1'].font = Font(name='Meiryo', bold=True, size=16, color='2F5496')

conditions = [
    ('分析目的', '日銀利上げサプライズでも下がりにくい大型株を抽出し、暴落リスクを限定する'),
    ('対象市場', '東証（プライム・スタンダード・グロース）'),
    ('時価総額フィルタ', '1,000億円以上'),
    ('流動性フィルタ', '直近25日平均売買代金 10億円/日以上'),
    ('対象銘柄数', f'{len(df)}銘柄'),
    ('ベンチマーク', 'TOPIX（1306.T ETFで代替）'),
    ('イベント窓', 't0（当日）、t0〜t+2（3営業日）、t0〜t+5（6営業日）'),
    ('Max Drawdown', '各窓内のピーク→ボトム最大下落率'),
    ('回帰調整', '相対リターンをUSD/JPY・S&P500・10年JGB ETFの変化で回帰し、残差を「純利上げ耐性」とする'),
    ('ランキング優先順位', '①ワースト回帰調整相対（高い順） → ②勝率 → ③平均回帰調整相対'),
    ('決算・IR除外', 'イベント当日/翌営業日に決算発表・大型IR重複→当該イベント回をN/A'),
    ('セクター分割', '金融（銀行・保険）と非金融を分離'),
    ('データ出所', 'Yahoo Finance（yfinance）、JPX東証上場銘柄一覧、OIS織り込みは各種報道'),
]

r = 3
ws.merge_cells(f'A{r}:B{r}')
ws[f'A{r}'] = '分析条件'
ws[f'A{r}'].font = Font(name='Meiryo', bold=True, size=12, color='2F5496')
r += 1
for label, val in conditions:
    ws[f'A{r}'] = label; ws[f'A{r}'].font = Font(name='Meiryo', bold=True, size=10)
    ws[f'B{r}'] = val; ws[f'B{r}'].font = nf
    r += 1

r += 1
ws.merge_cells(f'A{r}:B{r}')
ws[f'A{r}'] = 'イベント一覧（利上げサプライズ推定含む）'
ws[f'A{r}'].font = Font(name='Meiryo', bold=True, size=12, color='2F5496')
r += 1

evt_h = ['イベント', '発表日', '実際の利上げ(bp)', '事前OIS織込確率', '予想利上げ(bp)', 'サプライズ(bp)', 'サプライズ度合い']
for j, h in enumerate(evt_h):
    write_header(ws, r, j+1, h)
r += 1

for evt_name, evt_info in events.items():
    surprise_label = 'サプライズ大' if evt_info['surprise_bp'] >= 5 else ('小サプライズ' if evt_info['surprise_bp'] >= 2 else '織込済')
    vals = [evt_info['name'], evt_info['date'], evt_info['actual_hike_bp'],
            f"{evt_info['ois_probability']*100:.0f}%", round(evt_info['expected_hike_bp'], 1),
            round(evt_info['surprise_bp'], 1), surprise_label]
    for j, v in enumerate(vals):
        write_cell(ws, r, j+1, v)
    r += 1

r += 1
ws.merge_cells(f'A{r}:B{r}')
ws[f'A{r}'] = '回帰モデル係数（プール推定、t0〜t+2窓）'
ws[f'A{r}'].font = Font(name='Meiryo', bold=True, size=12, color='2F5496')
r += 1
reg_h = ['窓', 'N', 'R²', 'β(USD/JPY)', 'β(S&P500)', 'β(10Y JGB ETF)', '解釈']
for j, h in enumerate(reg_h):
    write_header(ws, r, j+1, h)
r += 1
reg_info = [
    ('t0', 1461, 0.0073, -0.519, -0.266, -0.027, '円高(USD/JPY↓)→株安の影響を除去'),
    ('t0〜t+2', 1461, 0.0030, 0.067, -0.216, 0.033, '米株連動を除去'),
    ('t0〜t+5', 1461, 0.0002, 0.091, -0.073, 0.018, '5日窓ではマクロ影響は希薄'),
]
for vals in reg_info:
    for j, v in enumerate(vals):
        write_cell(ws, r, j+1, v if not isinstance(v, float) else round(v, 4))
    r += 1

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 75
for c in 'CDEFGH':
    ws.column_dimensions[c].width = 18

# ==========================================
# Sheet 2: Macro (マクロ変数)
# ==========================================
ws_m = wb.create_sheet('Macro')
ws_m.merge_cells('A1:H1')
ws_m['A1'] = 'マクロ変数の変化（各イベント窓）'
ws_m['A1'].font = Font(name='Meiryo', bold=True, size=14, color='2F5496')

macro_h = ['イベント', '窓', 'TOPIX(%)', 'USD/JPY(%)', 'S&P500(%)', '10Y JGB ETF(%)', 'サプライズ(bp)']
for j, h in enumerate(macro_h):
    write_header(ws_m, 3, j+1, h)

r = 4
for evt_name, evt_info in events.items():
    for window, wlabel in [('t0', 't0'), ('t2', 't0〜t+2'), ('t5', 't0〜t+5')]:
        # Get macro values from first row (they're the same for all stocks)
        row0 = df.iloc[0]
        vals = [
            evt_info['name'], wlabel,
            round(row0.get(f'{evt_name}_topix_{window}', 0), 2),
            round(row0.get(f'{evt_name}_usdjpy_{window}', 0), 2),
            round(row0.get(f'{evt_name}_sp500_{window}', 0), 2),
            round(row0.get(f'{evt_name}_jgb_{window}', 0), 2),
            evt_info['surprise_bp'],
        ]
        for j, v in enumerate(vals):
            write_cell(ws_m, r, j+1, v)
        r += 1

for c in 'ABCDEFGH':
    ws_m.column_dimensions[c].width = 16

# ==========================================
# Sheet 3: EventReturns (全銘柄×全イベント×全窓)
# ==========================================
ws_e = wb.create_sheet('EventReturns')

# Build headers
base_cols = ['コード', '銘柄名', '業種', '時価総額(億円)', 'セクター']
evt_cols = []
for evt_name, evt_info in events.items():
    prefix = evt_info['name']
    for wlabel in ['t0', 't0~t+2', 't0~t+5']:
        evt_cols.extend([
            f'{prefix}\n{wlabel}\n銘柄(%)',
            f'{prefix}\n{wlabel}\nTOPIX(%)',
            f'{prefix}\n{wlabel}\n相対(%)',
            f'{prefix}\n{wlabel}\n回帰調整(%)',
            f'{prefix}\n{wlabel}\nMDD(%)',
        ])
    evt_cols.append(f'{prefix}\n決算/IR')

all_cols = base_cols + evt_cols + ['データ出所']

for j, h in enumerate(all_cols):
    write_header(ws_e, 1, j+1, h)

# Data rows
for i, (_, row) in enumerate(df.iterrows()):
    r = i + 2
    write_cell(ws_e, r, 1, row['code'])
    write_cell(ws_e, r, 2, row['name'])
    write_cell(ws_e, r, 3, row['sector_33'])
    write_cell(ws_e, r, 4, round(row['market_cap_100m'], 0) if pd.notna(row['market_cap_100m']) else '')
    write_cell(ws_e, r, 5, '金融' if row['is_financial'] else '非金融')
    
    col = 6
    for evt_name in ['event1', 'event2', 'event3']:
        for window, _ in [('t0', 't0'), ('t2', 't0~t+2'), ('t5', 't0~t+5')]:
            # Stock return
            v = row.get(f'{evt_name}_ret_{window}')
            write_cell(ws_e, r, col, round(v, 2) if pd.notna(v) else 'N/A', '0.00')
            col += 1
            # TOPIX return
            v = row.get(f'{evt_name}_topix_{window}')
            write_cell(ws_e, r, col, round(v, 2) if pd.notna(v) else '', '0.00')
            col += 1
            # Relative return
            v = row.get(f'{evt_name}_rel_{window}')
            c = write_cell(ws_e, r, col, round(v, 2) if pd.notna(v) else 'N/A', '0.00', color_val(v, -2, 2))
            col += 1
            # Adjusted relative
            v = row.get(f'{evt_name}_adj_rel_{window}')
            c = write_cell(ws_e, r, col, round(v, 2) if pd.notna(v) else 'N/A', '0.00', color_val(v, -2, 2))
            col += 1
            # MDD
            v = row.get(f'{evt_name}_mdd_{window}')
            c = write_cell(ws_e, r, col, round(v, 2) if pd.notna(v) else 'N/A', '0.00', color_val(v, -5, 0))
            col += 1
        
        # IR flag
        flag = row.get(f'{evt_name}_ir_flag', False)
        write_cell(ws_e, r, col, '●' if flag else '', fill=yellow_fill if flag else None)
        col += 1
    
    write_cell(ws_e, r, col, 'Yahoo Finance')

# Column widths
ws_e.column_dimensions['A'].width = 8
ws_e.column_dimensions['B'].width = 22
ws_e.column_dimensions['C'].width = 14
ws_e.column_dimensions['D'].width = 12
ws_e.column_dimensions['E'].width = 8
for j in range(6, len(all_cols) + 1):
    ws_e.column_dimensions[get_column_letter(j)].width = 12
ws_e.freeze_panes = 'F2'

# ==========================================
# Sheet 4 & 5: Summary_Financial / Summary_NonFinancial
# ==========================================
def write_summary(ws, ranking_df, title):
    ws.merge_cells('A1:P1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Meiryo', bold=True, size=14, color='2F5496')
    
    headers = [
        '順位', 'コード', '銘柄名', '業種', '時価総額\n(億円)',
        '有効\nデータ', '勝率\n(t0~t+2)',
        'ワースト\n回帰調整(%)', '平均\n回帰調整(%)',
        'ワースト\nMDD(%)',
        f'Event1\n2024/7/31\n回帰調整(%)', f'Event2\n2025/1/24\n回帰調整(%)', f'Event3\n2025/12/19\n回帰調整(%)',
        f'Event1\nMDD(%)', f'Event2\nMDD(%)', f'Event3\nMDD(%)',
        '決算/IR\nフラグ',
    ]
    
    for j, h in enumerate(headers):
        write_header(ws, 3, j+1, h)
    
    for i, (_, row) in enumerate(ranking_df.iterrows()):
        r = i + 4
        write_cell(ws, r, 1, int(row['rank']))
        write_cell(ws, r, 2, row['code'])
        c = write_cell(ws, r, 3, row['name'])
        c.alignment = Alignment(horizontal='left')
        c = write_cell(ws, r, 4, row['sector_33'])
        c.alignment = Alignment(horizontal='left')
        write_cell(ws, r, 5, round(row['market_cap_100m'], 0) if pd.notna(row['market_cap_100m']) else '')
        write_cell(ws, r, 6, int(row['valid_n_t2']))
        write_cell(ws, r, 7, f"{int(row['win_n_t2'])}/{int(row['valid_n_t2'])}")
        
        v = row['worst_adj_t2']
        write_cell(ws, r, 8, round(v, 2) if pd.notna(v) else '', '0.00', color_val(v, -1, 0))
        v = row['avg_adj_t2']
        write_cell(ws, r, 9, round(v, 2) if pd.notna(v) else '', '0.00')
        v = row['worst_mdd_t2']
        write_cell(ws, r, 10, round(v, 2) if pd.notna(v) else '', '0.00', color_val(v, -10, -3))
        
        for k, evt_name in enumerate(['event1', 'event2', 'event3']):
            v = row.get(f'{evt_name}_adj_rel_t2_excl')
            write_cell(ws, r, 11+k, round(v, 2) if pd.notna(v) else 'N/A', '0.00', color_val(v, -2, 2))
        
        for k, evt_name in enumerate(['event1', 'event2', 'event3']):
            v = row.get(f'{evt_name}_mdd_t2_excl')
            write_cell(ws, r, 14+k, round(v, 2) if pd.notna(v) else 'N/A', '0.00', color_val(v, -10, -3))
        
        flags = []
        for evt_name, evt_info in events.items():
            if row.get(f'{evt_name}_ir_flag', False):
                flags.append(evt_info['name'][:7])
        write_cell(ws, r, 17, ', '.join(flags) if flags else '', fill=yellow_fill if flags else None)
    
    widths = [6, 8, 24, 14, 12, 6, 8, 12, 12, 12, 12, 12, 12, 10, 10, 10, 16]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(j+1)].width = w
    ws.freeze_panes = 'D4'

ws_fin = wb.create_sheet('Summary_Financial')
write_summary(ws_fin, fin_rank, '金融セクター 利上げ耐性ランキング（回帰調整版）')

ws_nf = wb.create_sheet('Summary_NonFinancial')
write_summary(ws_nf, nonfin_rank, '非金融セクター 利上げ耐性ランキング（回帰調整版）')

# Save
out = '/home/ubuntu/boj_analysis/BOJ_RateHike_Analysis_V2.xlsx'
wb.save(out)
print(f"Excel saved: {out}")
print(f"Sheets: {wb.sheetnames}")
